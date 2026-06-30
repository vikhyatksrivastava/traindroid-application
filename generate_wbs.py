"""
traindroid WBS Generator  v2.0
Produces traindroid_WBS.xlsx with per-discipline estimates.
Includes: UX Design, Architecture Design, Observability, Documentation,
and all 4 core differentiator features explicitly tagged.
"""

from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import (
    PatternFill, Font, Alignment, Border, Side
)
from openpyxl.utils import get_column_letter

# ──────────────────────────────────────────────────────────────────────────────
# DATA — WBS rows
# Columns: id, phase, feature, sub_feature, tag, fe, be, qa, devops, db
#
# Tag values:
#   "★★★ Core Diff"  — one of the 4 platform differentiators
#   "★★  MVP"        — required for Milestone 1 working prototype
#   "★    M2+"       — Milestone 2 or later
#   "—   Design"     — pre-build design / discovery activity
# ──────────────────────────────────────────────────────────────────────────────

T_CORE  = "★★★ Core Diff"
T_MVP   = "★★  MVP"
T_M2    = "★    M2+"
T_DSGN  = "—   Design"

WBS = [
    # ── Phase D — UX & Product Design ────────────────────────────────────────
    ("D.1",  "Phase D — UX & Product Design",     "User Research & Personas",            "Identify QA engineer, super-admin, and CI-trigger user journeys",     T_DSGN,  4,  0,  2,  0,  0),
    ("D.2",  "Phase D — UX & Product Design",     "Information Architecture",            "Site map, screen inventory, navigation hierarchy",                    T_DSGN,  4,  0,  0,  0,  0),
    ("D.3",  "Phase D — UX & Product Design",     "Low-Fidelity Wireframes",             "All 6 portal screens: upload, config, progress, results, audit, admin",T_DSGN,  8,  0,  2,  0,  0),
    ("D.4",  "Phase D — UX & Product Design",     "High-Fidelity Mockups",               "Figma mockups with full colour, typography, spacing, component spec", T_DSGN, 12,  0,  2,  0,  0),
    ("D.5",  "Phase D — UX & Product Design",     "Design System / Component Library",   "Tokens: colour, type, spacing; Button, Card, Table, Badge components", T_DSGN,  8,  0,  1,  0,  0),
    ("D.6",  "Phase D — UX & Product Design",     "Interactive Prototype",               "Clickable Figma prototype for stakeholder sign-off walkthrough",       T_DSGN,  4,  0,  1,  0,  0),
    ("D.7",  "Phase D — UX & Product Design",     "UX Review & Sign-Off",                "Stakeholder walkthrough, feedback incorporation, final approval",      T_DSGN,  2,  0,  1,  0,  0),
    ("D.8",  "Phase D — UX & Product Design",     "Accessibility (a11y) Design Review",  "WCAG 2.1 AA check on mockups: contrast, focus order, ARIA labels",    T_DSGN,  2,  0,  2,  0,  0),

    # ── Phase A — Architecture & Technical Design ─────────────────────────────
    ("A.1",  "Phase A — Architecture Design",     "System Architecture Diagram",         "C4 context + container diagram: portal, runner, emulator, LLM, DB",   T_DSGN,  0,  4,  0,  2,  0),
    ("A.2",  "Phase A — Architecture Design",     "API Contract Design (OpenAPI)",       "OpenAPI 3.1 spec for all endpoints; Swagger UI hosted on portal",      T_DSGN,  0,  6,  1,  0,  0),
    ("A.3",  "Phase A — Architecture Design",     "Data Flow Diagrams",                  "APK upload → queue → runner → LLM → audit → approval flow",           T_DSGN,  0,  3,  0,  1,  0),
    ("A.4",  "Phase A — Architecture Design",     "Database ER Diagram",                 "Tables: users, runs, steps, screenshots, reports, approvals",          T_DSGN,  0,  2,  0,  0,  4),
    ("A.5",  "Phase A — Architecture Design",     "Architecture Decision Records (ADR)", "ADR for: LLM provider, DB choice, queue strategy, Docker vs bare metal",T_DSGN,  0,  4,  0,  1,  0),
    ("A.6",  "Phase A — Architecture Design",     "Technology Spike — Emulator in Docker","PoC: headless AVD boots inside Docker on local machine (ARM + x86)",  T_DSGN,  0,  0,  2, 12,  0),
    ("A.7",  "Phase A — Architecture Design",     "Technology Spike — LLM Action Schema","PoC: structured JSON action output from Claude Haiku via prompt eng.", T_DSGN,  0,  6,  2,  0,  0),
    ("A.8",  "Phase A — Architecture Design",     "Security Threat Model",               "STRIDE model for file upload, admin auth, and AI output trust boundary",T_DSGN,  0,  3,  2,  0,  0),

    # ── Phase 0 — Project Setup ───────────────────────────────────────────────
    ("0.1",  "Phase 0 — Project Setup",           "Repository & Monorepo Setup",         "Git repo, folder structure, .gitignore, branch strategy, README",     T_MVP,   0,  2,  0,  4,  0),
    ("0.2",  "Phase 0 — Project Setup",           "Docker Compose Skeleton",             "Node portal + Python test runner containers, networks, volumes",       T_MVP,   0,  2,  1,  6,  0),
    ("0.3",  "Phase 0 — Project Setup",           "Linting & Formatting Config",         "ESLint (JS), Black/Ruff (Python), pre-commit hooks",                  T_MVP,   1,  2,  0,  1,  0),
    ("0.4",  "Phase 0 — Project Setup",           "Environment Config & Secrets",        ".env handling, dotenv validation, secret rotation strategy",          T_MVP,   0,  2,  0,  2,  0),
    ("0.5",  "Phase 0 — Project Setup",           "Developer Environment Setup Guide",   "Step-by-step local dev setup; verify with fresh-clone checklist",      T_MVP,   0,  2,  0,  1,  0),
    ("0.6",  "Phase 0 — Project Setup",           "Database Schema & Migrations",        "Initial schema SQL, migration tool (Flyway / Alembic), seed data",    T_MVP,   0,  2,  0,  0,  6),
    ("0.7",  "Phase 0 — Project Setup",           "Job Queue Setup",                     "Bull/BullMQ or Python RQ for async test run jobs",                    T_MVP,   0,  4,  1,  2,  0),

    # ── Phase 1 — APK Upload & Self-Hosted Pipeline  [CORE DIFF #1] ──────────
    ("1.1",  "Phase 1 — APK Upload & Self-Hosted", "APK Upload Endpoint",                "POST /upload — multipart form, file save, job enqueue",               T_CORE,  0,  4,  2,  0,  0),
    ("1.2",  "Phase 1 — APK Upload & Self-Hosted", "Self-Hosted Local Volume Storage",   "Disk path resolver, SHA-256 filename hash, directory isolation",       T_CORE,  0,  3,  1,  2,  2),
    ("1.3",  "Phase 1 — APK Upload & Self-Hosted", "File Validation & Security Checks",  "MIME type, magic bytes, size cap, APK ZIP header check",              T_CORE,  0,  3,  2,  0,  0),
    ("1.4",  "Phase 1 — APK Upload & Self-Hosted", "APK Retention & Cleanup Policy",     "TTL-based deletion, manual purge endpoint, disk-space guard",         T_MVP,   0,  2,  1,  0,  2),
    ("1.5",  "Phase 1 — APK Upload & Self-Hosted", "Android Emulator Docker Image",      "AVD creation, headless boot (-no-window), KVM/x86 emulation",         T_CORE,  0,  0,  2, 12,  0),
    ("1.6",  "Phase 1 — APK Upload & Self-Hosted", "ADB Connectivity Layer",             "adb connect, device ready poll, install APK, launch app, teardown",   T_CORE,  0,  6,  2,  2,  0),
    ("1.7",  "Phase 1 — APK Upload & Self-Hosted", "Emulator Health Check & Pre-warm",   "Boot readiness polling, keep-alive between runs, restart on crash",   T_MVP,   0,  4,  1,  2,  0),
    ("1.8",  "Phase 1 — APK Upload & Self-Hosted", "Emulator Pool Manager",              "Queue-based slot allocation, single emulator for MVP, extendable",    T_M2,    0,  6,  2,  2,  0),

    # ── Phase 2 — Zero-Script AI Agent Engine  [CORE DIFF #2] ────────────────
    ("2.1",  "Phase 2 — AI Agent Engine",         "Screenshot Capture",                  "adb screencap → PNG → base64 encode → attach to LLM payload",         T_CORE,  0,  3,  1,  0,  0),
    ("2.2",  "Phase 2 — AI Agent Engine",         "UI Accessibility Tree Dump",          "uiautomator dump → XML parse → structured node dict",                 T_CORE,  0,  4,  2,  0,  0),
    ("2.3",  "Phase 2 — AI Agent Engine",         "LLM Prompt Engineering & Iteration",  "System prompt design, few-shot examples, JSON schema enforcement",    T_CORE,  0,  8,  2,  0,  0),
    ("2.4",  "Phase 2 — AI Agent Engine",         "LLM Integration — Tier 1 (text)",     "Claude Haiku / GPT-4o-mini: accessibility tree → action JSON",        T_CORE,  0,  6,  3,  0,  0),
    ("2.5",  "Phase 2 — AI Agent Engine",         "LLM Integration — Tier 2 (vision)",   "Claude Sonnet / GPT-4o: screenshot + tree → action JSON",             T_CORE,  0,  6,  3,  0,  0),
    ("2.6",  "Phase 2 — AI Agent Engine",         "Action Response Schema Design",        "Pydantic/Zod schema: action_type, target, value, confidence, reason", T_CORE,  0,  4,  2,  0,  0),
    ("2.7",  "Phase 2 — AI Agent Engine",         "Action Execution via ADB",             "tap, swipe, long_press, type, back, scroll — ADB command mapper",    T_CORE,  0,  6,  3,  0,  0),
    ("2.8",  "Phase 2 — AI Agent Engine",         "Goal / Flow Parsing",                  "Plain-English goal → sub-goals + success criteria via LLM",          T_CORE,  0,  6,  2,  0,  0),
    ("2.9",  "Phase 2 — AI Agent Engine",         "Agent Loop Orchestrator",              "Step driver: capture → reason → act → log → next step",              T_CORE,  0, 10,  4,  0,  0),
    ("2.10", "Phase 2 — AI Agent Engine",         "Tiered Model Routing Logic",           "Tier 1 first; escalate to Tier 2 on low-confidence flag",            T_CORE,  0,  4,  2,  0,  0),
    ("2.11", "Phase 2 — AI Agent Engine",         "Confidence Threshold & Review Flag",   "Score threshold → pause run → flag step for human review",           T_CORE,  0,  4,  2,  0,  0),
    ("2.12", "Phase 2 — AI Agent Engine",         "Stuck-Screen Detection",               "Pixel-hash comparison; max-steps guard; graceful abort + report",    T_MVP,   0,  4,  2,  0,  0),
    ("2.13", "Phase 2 — AI Agent Engine",         "AI Cost Tracking per Run",             "Token counter middleware; log Tier 1 vs Tier 2 usage per run",        T_M2,    0,  3,  1,  0,  2),

    # ── Phase 3 — Screen Flow Testing ────────────────────────────────────────
    ("3.1",  "Phase 3 — Screen Flow Testing",     "Flow Definition Format (JSON/YAML)",   "Schema: steps, expected outcomes, branch conditions, success signal", T_MVP,   0,  3,  1,  0,  2),
    ("3.2",  "Phase 3 — Screen Flow Testing",     "Flow Execution Engine",                "Step-by-step runner, pre/post conditions, conditional branching",     T_MVP,   0,  8,  3,  0,  0),
    ("3.3",  "Phase 3 — Screen Flow Testing",     "Pass/Fail Detection Per Step",         "Element presence, text match, screen-hash compare, timeout guard",    T_MVP,   0,  4,  3,  0,  0),
    ("3.4",  "Phase 3 — Screen Flow Testing",     "Screenshot Capture on Failure",        "Auto-capture + annotate (timestamp overlay) on any step failure",     T_MVP,   0,  2,  1,  0,  0),
    ("3.5",  "Phase 3 — Screen Flow Testing",     "Signup + Login Flow (Reference Flow)", "Built-in flow: detect create-account, fill dummy creds, verify login", T_MVP,   0,  4,  2,  0,  0),

    # ── Phase 4 — Functionality Testing ──────────────────────────────────────
    ("4.1",  "Phase 4 — Functionality Testing",   "Input Simulation",                     "Tap, swipe, long-press, type — coordinate + accessibility target",   T_MVP,   0,  4,  2,  0,  0),
    ("4.2",  "Phase 4 — Functionality Testing",   "Assertion Engine",                     "Element presence, text value, enabled/disabled, visible/hidden",      T_MVP,   0,  6,  3,  0,  0),
    ("4.3",  "Phase 4 — Functionality Testing",   "PyTest Integration & Test Runner",      "Fixture setup, parametrised cases, JUnit XML output for CI",         T_MVP,   0,  4,  4,  0,  0),

    # ── Phase 5 — LLM Reasoning Audit Trail  [CORE DIFF #3] ──────────────────
    ("5.1",  "Phase 5 — Audit & LLM Reasoning",  "Per-Step Audit Log Model",             "Schema: timestamp, action, screenshot_id, status, llm_reasoning",    T_CORE,  0,  4,  2,  0,  2),
    ("5.2",  "Phase 5 — Audit & LLM Reasoning",  "LLM Reasoning Narrative Formatter",    "Convert raw LLM JSON → human-readable prose per step (e.g. 'Detected signup screen…')", T_CORE, 0, 6, 2, 0, 0),
    ("5.3",  "Phase 5 — Audit & LLM Reasoning",  "Screenshot Storage & Retrieval",       "Save screenshots to volume, index by run+step, serve via API",        T_CORE,  0,  3,  1,  1,  2),
    ("5.4",  "Phase 5 — Audit & LLM Reasoning",  "Test Result Aggregation Service",      "Roll up step results → run-level pass/fail + failure summary",        T_CORE,  0,  6,  2,  0,  2),
    ("5.5",  "Phase 5 — Audit & LLM Reasoning",  "Audit Trail Immutability",             "Append-only log writes; no update/delete on audit rows post-run",     T_CORE,  0,  2,  1,  0,  2),
    ("5.6",  "Phase 5 — Audit & LLM Reasoning",  "HTML Report Generation",               "Jinja2 template: step table, embedded screenshots, reasoning prose",  T_MVP,   2,  4,  2,  0,  0),
    ("5.7",  "Phase 5 — Audit & LLM Reasoning",  "PDF Report Export",                    "WeasyPrint render of HTML report; download via portal",               T_MVP,   1,  3,  1,  0,  0),
    ("5.8",  "Phase 5 — Audit & LLM Reasoning",  "Report Storage & Retrieval API",       "Persist report file, GET /report/:id, serve on approval",             T_MVP,   0,  3,  1,  1,  0),

    # ── Phase 6 — Web Portal (Frontend) ──────────────────────────────────────
    ("6.1",  "Phase 6 — Web Portal",              "App Shell, Routing & Auth Layout",     "React/Next.js scaffold, protected routes, responsive nav sidebar",    T_MVP,   8,  2,  2,  0,  0),
    ("6.2",  "Phase 6 — Web Portal",              "Design System Implementation",         "Translate Figma tokens → Tailwind/CSS vars; shared component lib",    T_MVP,   8,  0,  2,  0,  0),
    ("6.3",  "Phase 6 — Web Portal",              "APK Upload Screen",                    "Drag-and-drop zone, progress bar, file info, validation feedback",    T_CORE,  8,  2,  2,  0,  0),
    ("6.4",  "Phase 6 — Web Portal",              "Test Run Configuration Screen",        "Goal textarea, model tier selector, run name/tag, submit CTA",        T_MVP,   6,  2,  2,  0,  0),
    ("6.5",  "Phase 6 — Web Portal",              "Live Test Run Progress View",          "WebSocket feed, step timeline, live screenshot carousel, spinner",    T_MVP,  10,  4,  3,  0,  0),
    ("6.6",  "Phase 6 — Web Portal",              "Results / Audit Review Screen",        "Step-by-step audit table, LLM reasoning prose, pass/fail badges",     T_CORE,  8,  2,  3,  0,  0),
    ("6.7",  "Phase 6 — Web Portal",              "Admin Approval Screen",                "Pending runs list, approve/reject, sign-off timestamp, notes input",  T_CORE,  6,  2,  2,  0,  0),
    ("6.8",  "Phase 6 — Web Portal",              "Loading, Error & Empty States",        "Skeleton loaders, error banners, empty-state illustrations per screen",T_MVP,   4,  0,  2,  0,  0),
    ("6.9",  "Phase 6 — Web Portal",              "Accessibility (a11y) Implementation",  "Keyboard navigation, ARIA labels, focus management, contrast audit",  T_M2,    4,  0,  3,  0,  0),

    # ── Phase 7 — Admin Approval Workflow  [CORE DIFF #4] ────────────────────
    ("7.1",  "Phase 7 — Admin Approval Workflow", "Super Admin User Model & Auth",        "User table, role enum, JWT login/refresh, bcrypt password hashing",   T_CORE,  1,  4,  2,  0,  3),
    ("7.2",  "Phase 7 — Admin Approval Workflow", "Approve / Reject API Endpoint",        "PATCH /run/:id/approve|reject, auth guard, state machine transition", T_CORE,  0,  4,  2,  0,  0),
    ("7.3",  "Phase 7 — Admin Approval Workflow", "Report Release Gating",                "Report file only served / generated after admin approval action",     T_CORE,  0,  3,  2,  0,  0),
    ("7.4",  "Phase 7 — Admin Approval Workflow", "Audit Sign-off Recording",             "Immutable record: approver, timestamp, admin notes, run_id FK",       T_CORE,  0,  2,  1,  0,  2),
    ("7.5",  "Phase 7 — Admin Approval Workflow", "Admin Notification (Email)",           "Send email on new run awaiting approval; configurable SMTP",          T_M2,    0,  3,  1,  0,  0),

    # ── Phase 8 — CI/CD Integration ───────────────────────────────────────────
    ("8.1",  "Phase 8 — CI/CD Integration",       "GitHub Actions Workflow",              "Trigger test run on push/PR; poll for result; fail CI on test fail",  T_M2,    0,  2,  1,  6,  0),
    ("8.2",  "Phase 8 — CI/CD Integration",       "Webhook / API Trigger Endpoint",       "POST /trigger; HMAC-SHA256 signature validation; job enqueue",        T_M2,    0,  4,  2,  2,  0),
    ("8.3",  "Phase 8 — CI/CD Integration",       "Status Badge & Completion Callback",   "Badge SVG endpoint; POST callback URL on run completion",             T_M2,    0,  3,  1,  2,  0),
    ("8.4",  "Phase 8 — CI/CD Integration",       "Docker Multi-Stage Build",             "Slim production image, layer caching, build-arg injection",           T_M2,    0,  0,  1,  6,  0),

    # ── Phase 9 — Observability & Operations (NEW) ────────────────────────────
    ("9.1",  "Phase 9 — Observability",           "Structured Logging (Winston/Pino)",    "JSON logs with run_id, step_id, level, duration on every service",    T_MVP,   0,  4,  1,  0,  0),
    ("9.2",  "Phase 9 — Observability",           "Error Tracking (Sentry)",              "Sentry SDK in Node + Python; alert on unhandled exceptions",          T_MVP,   0,  2,  1,  2,  0),
    ("9.3",  "Phase 9 — Observability",           "Health Check Endpoints",               "GET /health on portal + runner; Docker HEALTHCHECK integration",      T_MVP,   0,  2,  1,  1,  0),
    ("9.4",  "Phase 9 — Observability",           "Run Failure Alerting",                 "Slack/email alert when a test run fails or gets stuck",               T_M2,    0,  3,  1,  0,  0),
    ("9.5",  "Phase 9 — Observability",           "Basic Metrics Dashboard",              "Grafana + Prometheus: run count, success rate, avg duration",         T_M2,    0,  0,  1,  6,  0),
    ("9.6",  "Phase 9 — Observability",           "LLM API Cost Monitoring",              "Per-run token usage logged; daily cost summary report",               T_M2,    0,  2,  1,  0,  2),

    # ── Phase 10 — Documentation (NEW) ───────────────────────────────────────
    ("10.1", "Phase 10 — Documentation",          "Developer Setup Guide",                "Clone → configure → run locally in < 15 min; tested on clean machine",T_MVP,   0,  3,  1,  1,  0),
    ("10.2", "Phase 10 — Documentation",          "API Reference (Swagger UI)",           "Auto-generated from OpenAPI spec; hosted at /api/docs",               T_MVP,   0,  2,  1,  0,  0),
    ("10.3", "Phase 10 — Documentation",          "User Portal Guide",                    "How to upload APK, configure a run, read audit report, approve",      T_MVP,   2,  0,  1,  0,  0),
    ("10.4", "Phase 10 — Documentation",          "Deployment Guide (Self-Hosted)",       "Docker Compose prod setup, env vars, volume mounts, KVM prereqs",     T_MVP,   0,  3,  0,  2,  0),
    ("10.5", "Phase 10 — Documentation",          "CONTRIBUTING.md & Coding Standards",   "PR process, branch naming, commit conventions, review checklist",     T_M2,    0,  2,  0,  0,  0),
    ("10.6", "Phase 10 — Documentation",          "Architecture Overview Document",        "Written summary of C4 diagrams, data flow, design decisions",         T_M2,    0,  3,  0,  1,  0),

    # ── Phase 11 — Platform QA ────────────────────────────────────────────────
    ("11.1", "Phase 11 — Platform QA",            "Unit Tests — Backend API",             "Jest/Pytest unit tests for all endpoints, helpers, schemas",          T_MVP,   0,  2,  8,  0,  0),
    ("11.2", "Phase 11 — Platform QA",            "Integration Tests — Emulator Pipeline","Full pipeline test with a sample APK; assert audit log output",       T_MVP,   0,  2,  8,  0,  0),
    ("11.3", "Phase 11 — Platform QA",            "End-to-End Portal Tests (Playwright)", "E2E: upload APK → run → live view → results → approve flow",         T_M2,    2,  0, 10,  0,  0),
    ("11.4", "Phase 11 — Platform QA",            "Performance & Load Testing",           "Concurrent run stress test, emulator queue saturation",               T_M2,    0,  0,  6,  2,  0),
    ("11.5", "Phase 11 — Platform QA",            "Security Testing (OWASP Top 10)",      "File upload checks, auth bypass, injection probes, JWT fuzz",         T_MVP,   0,  2,  6,  0,  0),
    ("11.6", "Phase 11 — Platform QA",            "AI Output Validation Testing",         "Inject known screens; verify LLM action accuracy + reasoning prose",  T_MVP,   0,  2,  6,  0,  0),
]

# ──────────────────────────────────────────────────────────────────────────────
# COLOUR PALETTE
# ──────────────────────────────────────────────────────────────────────────────

PHASE_COLORS = {
    "Phase D — UX & Product Design":      "6E2594",   # violet
    "Phase A — Architecture Design":      "2D6A4F",   # dark teal
    "Phase 0 — Project Setup":            "1F3864",   # dark navy
    "Phase 1 — APK Upload & Self-Hosted": "1D6A96",   # teal-blue
    "Phase 2 — AI Agent Engine":          "375623",   # forest green
    "Phase 3 — Screen Flow Testing":      "843C0C",   # burnt orange
    "Phase 4 — Functionality Testing":    "4A235A",   # deep purple
    "Phase 5 — Audit & LLM Reasoning":   "7B3F00",   # dark brown
    "Phase 6 — Web Portal":               "154360",   # dark blue
    "Phase 7 — Admin Approval Workflow":  "512E5F",   # plum
    "Phase 8 — CI/CD Integration":        "1A5276",   # slate blue
    "Phase 9 — Observability":            "0B3D0B",   # deep green
    "Phase 10 — Documentation":           "4D4D00",   # olive
    "Phase 11 — Platform QA":             "1B4F72",   # steel blue
}

PHASE_LIGHT = {
    "Phase D — UX & Product Design":      "F3E5F5",
    "Phase A — Architecture Design":      "D7F5E9",
    "Phase 0 — Project Setup":            "D6E4F0",
    "Phase 1 — APK Upload & Self-Hosted": "D6EAF8",
    "Phase 2 — AI Agent Engine":          "D5F5E3",
    "Phase 3 — Screen Flow Testing":      "FDEBD0",
    "Phase 4 — Functionality Testing":    "F5EEF8",
    "Phase 5 — Audit & LLM Reasoning":   "FBEEE6",
    "Phase 6 — Web Portal":               "D6EAF8",
    "Phase 7 — Admin Approval Workflow":  "F4ECF7",
    "Phase 8 — CI/CD Integration":        "D6EAF8",
    "Phase 9 — Observability":            "E9F7EF",
    "Phase 10 — Documentation":           "FAFAD2",
    "Phase 11 — Platform QA":             "EBF5FB",
}

# Tag colours (background fill for Tag column)
TAG_COLORS = {
    T_CORE:  ("C00000", "FFFFFF"),   # red bg, white text
    T_MVP:   ("375623", "FFFFFF"),   # green bg, white text
    T_M2:    ("44546A", "FFFFFF"),   # grey-blue bg, white text
    T_DSGN:  ("6E2594", "FFFFFF"),   # violet bg, white text
}

# Discipline highlight colours
COL_COLORS = {
    "Frontend":  "2E75B6",
    "Backend":   "375623",
    "QA":        "833C00",
    "DevOps":    "44546A",
    "Database":  "7030A0",
    "Total":     "1F3864",
}


def thin_border():
    s = Side(style="thin", color="CCCCCC")
    return Border(left=s, right=s, top=s, bottom=s)


def header_border():
    s = Side(style="medium", color="FFFFFF")
    return Border(left=s, right=s, top=s, bottom=s)


def build_wbs_sheet(wb):
    # ── Column layout (11 cols) ───────────────────────────────────────────────
    # A=ID  B=Phase  C=Feature  D=Sub-Feature  E=Tag
    # F=FE  G=BE  H=QA  I=DevOps  J=DB  K=Total
    ws = wb.active
    ws.title = "WBS — Full Breakdown"
    ws.sheet_view.showGridLines = False

    col_widths = [6, 28, 26, 52, 18, 11, 11, 11, 11, 11, 11]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    ws.row_dimensions[1].height = 20
    ws.row_dimensions[2].height = 30
    ws.row_dimensions[3].height = 26

    # ── Row 1 — project title ─────────────────────────────────────────────────
    ws.merge_cells("A1:K1")
    t = ws["A1"]
    t.value = "traindroid — Autonomous Android UI Testing Platform"
    t.font = Font(name="Calibri", bold=True, size=16, color="FFFFFF")
    t.fill = PatternFill("solid", fgColor="1F3864")
    t.alignment = Alignment(horizontal="center", vertical="center")

    # ── Row 2 — subtitle ─────────────────────────────────────────────────────
    ws.merge_cells("A2:K2")
    sub = ws["A2"]
    sub.value = "Work Breakdown Structure & Prototype Estimate  ·  v2.0 · May 2026"
    sub.font = Font(name="Calibri", italic=True, size=11, color="D9E1F2")
    sub.fill = PatternFill("solid", fgColor="2F5496")
    sub.alignment = Alignment(horizontal="center", vertical="center")

    # ── Row 3 — legend ───────────────────────────────────────────────────────
    ws.merge_cells("A3:K3")
    leg = ws["A3"]
    leg.value = (
        "   Tag legend:  ★★★ Core Differentiator  |  ★★ MVP Required  |  ★ Milestone 2+  |  — Design / Discovery"
    )
    leg.font = Font(name="Calibri", bold=True, size=9, color="FFFFFF")
    leg.fill = PatternFill("solid", fgColor="44546A")
    leg.alignment = Alignment(horizontal="left", vertical="center", indent=1)

    # ── Row 4 — column headers ───────────────────────────────────────────────
    ws.row_dimensions[4].height = 28
    col_headers = [
        ("ID",                  "2F5496"),
        ("Phase",               "2F5496"),
        ("Feature",             "2F5496"),
        ("Sub-Feature / Deliverable", "2F5496"),
        ("Priority / Tag",      "44546A"),
        ("Frontend\n(hrs)",     COL_COLORS["Frontend"]),
        ("Backend\n(hrs)",      COL_COLORS["Backend"]),
        ("QA\n(hrs)",           COL_COLORS["QA"]),
        ("DevOps\n(hrs)",       COL_COLORS["DevOps"]),
        ("Database\n(hrs)",     COL_COLORS["Database"]),
        ("TOTAL\n(hrs)",        COL_COLORS["Total"]),
    ]
    for col, (header, color) in enumerate(col_headers, 1):
        cell = ws.cell(row=4, column=col, value=header)
        cell.fill = PatternFill("solid", fgColor=color)
        cell.font = Font(name="Calibri", bold=True, size=10, color="FFFFFF")
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = header_border()

    ws.freeze_panes = "A5"

    # ── Data rows ─────────────────────────────────────────────────────────────
    row = 5
    prev_phase = None
    phase_rows = {}

    for (id_, phase, feature, sub, tag, fe, be, qa, devops, db) in WBS:
        total = fe + be + qa + devops + db
        light = PHASE_LIGHT.get(phase, "FFFFFF")
        dark  = PHASE_COLORS.get(phase, "1F3864")

        if phase != prev_phase:
            ws.merge_cells(f"A{row}:K{row}")
            ph = ws.cell(row=row, column=1, value=f"  {phase}")
            ph.font = Font(name="Calibri", bold=True, size=11, color="FFFFFF")
            ph.fill = PatternFill("solid", fgColor=dark)
            ph.alignment = Alignment(vertical="center")
            ph.border = thin_border()
            ws.row_dimensions[row].height = 20
            row += 1
            prev_phase = phase

        phase_rows.setdefault(phase, []).append(row)

        values = [id_, phase, feature, sub, tag, fe, be, qa, devops, db, total]
        for col, val in enumerate(values, 1):
            cell = ws.cell(row=row, column=col, value=val)
            cell.fill = PatternFill("solid", fgColor=light)
            cell.border = thin_border()

            if col == 1:   # ID
                cell.font = Font(name="Calibri", bold=True, size=9, color=dark)
                cell.alignment = Alignment(horizontal="center", vertical="center")
            elif col == 2:  # Phase (hide — redundant under separator)
                cell.font = Font(name="Calibri", size=8, color="AAAAAA")
                cell.alignment = Alignment(vertical="center")
            elif col == 3:  # Feature
                cell.font = Font(name="Calibri", bold=True, size=9, color="222222")
                cell.alignment = Alignment(vertical="center", wrap_text=True)
            elif col == 4:  # Sub-feature
                cell.font = Font(name="Calibri", size=9, color="444444", italic=True)
                cell.alignment = Alignment(vertical="center", wrap_text=True)
            elif col == 5:  # Tag
                tag_bg, tag_fg = TAG_COLORS.get(tag, ("EEEEEE", "333333"))
                cell.fill = PatternFill("solid", fgColor=tag_bg)
                cell.font = Font(name="Calibri", bold=True, size=8, color=tag_fg)
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            elif col in (6, 7, 8, 9, 10):  # discipline hours
                disc = ["Frontend","Backend","QA","DevOps","Database"][col-6]
                cell.font = Font(name="Calibri", size=10, bold=(val > 0),
                                 color=(COL_COLORS[disc] if val > 0 else "CCCCCC"))
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.number_format = '0'
            elif col == 11:  # Total
                cell.font = Font(name="Calibri", bold=True, size=10, color=dark)
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.number_format = '0'

        ws.row_dimensions[row].height = 28
        row += 1

    # ── Phase subtotals ───────────────────────────────────────────────────────
    for phase, rows in phase_rows.items():
        dark = PHASE_COLORS.get(phase, "1F3864")
        ws.merge_cells(f"A{row}:E{row}")
        sc = ws.cell(row=row, column=1, value=f"  Subtotal — {phase}")
        sc.font = Font(name="Calibri", bold=True, size=9, color="FFFFFF")
        sc.fill = PatternFill("solid", fgColor=dark)
        sc.alignment = Alignment(vertical="center")
        sc.border = thin_border()

        for i, disc_key in enumerate(["Frontend","Backend","QA","DevOps","Database","Total"]):
            src_col = 6 + i
            total_val = sum(ws.cell(row=r, column=src_col).value or 0 for r in rows)
            c = ws.cell(row=row, column=src_col, value=total_val)
            c.fill = PatternFill("solid", fgColor=dark)
            c.font = Font(name="Calibri", bold=True, size=10, color="FFFFFF")
            c.alignment = Alignment(horizontal="center", vertical="center")
            c.border = header_border()
            c.number_format = '0'

        ws.row_dimensions[row].height = 20
        row += 1

    # ── Grand Total ───────────────────────────────────────────────────────────
    row += 1
    ws.merge_cells(f"A{row}:E{row}")
    gt = ws.cell(row=row, column=1, value="  GRAND TOTAL — PROTOTYPE BUILD")
    gt.font = Font(name="Calibri", bold=True, size=12, color="FFFFFF")
    gt.fill = PatternFill("solid", fgColor="1F3864")
    gt.alignment = Alignment(vertical="center")
    gt.border = header_border()

    all_fe  = sum(r[5] for r in WBS)
    all_be  = sum(r[6] for r in WBS)
    all_qa  = sum(r[7] for r in WBS)
    all_dv  = sum(r[8] for r in WBS)
    all_db  = sum(r[9] for r in WBS)
    all_tot = all_fe + all_be + all_qa + all_dv + all_db

    for i, (val, key) in enumerate(zip(
            [all_fe, all_be, all_qa, all_dv, all_db, all_tot],
            ["Frontend","Backend","QA","DevOps","Database","Total"])):
        cell = ws.cell(row=row, column=6+i, value=val)
        cell.fill = PatternFill("solid", fgColor=COL_COLORS[key])
        cell.font = Font(name="Calibri", bold=True, size=12, color="FFFFFF")
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = header_border()
        cell.number_format = '0'
    ws.row_dimensions[row].height = 28

    return wb, all_fe, all_be, all_qa, all_dv, all_db, all_tot


def build_summary_sheet(wb, all_fe, all_be, all_qa, all_dv, all_db, all_tot):
    ws = wb.create_sheet("Summary Dashboard")
    ws.sheet_view.showGridLines = False

    col_widths = [4, 32, 18, 18, 20, 4]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # Title
    ws.merge_cells("A1:F1")
    t = ws["A1"]
    t.value = "traindroid — Prototype Effort Summary"
    t.font = Font(name="Calibri", bold=True, size=15, color="FFFFFF")
    t.fill = PatternFill("solid", fgColor="1F3864")
    t.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 30

    ws.merge_cells("A2:F2")
    sub = ws["A2"]
    sub.value = "Total estimated effort for a working prototype  ·  v2.0 May 2026"
    sub.font = Font(name="Calibri", italic=True, size=10, color="888888")
    sub.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[2].height = 18

    # ── Core Differentiators banner ───────────────────────────────────────────
    ws.row_dimensions[4].height = 20
    ws.merge_cells("B4:F4")
    cd_hdr = ws["B4"]
    cd_hdr.value = "Core Platform Differentiators (★★★)"
    cd_hdr.font = Font(name="Calibri", bold=True, size=11, color="FFFFFF")
    cd_hdr.fill = PatternFill("solid", fgColor="C00000")
    cd_hdr.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    cd_hdr.border = thin_border()

    core_items = [
        ("#1", "APK Upload → Self-Hosted",          "Phase 1 — no cloud lock-in; local volume; Docker-first"),
        ("#2", "Zero-Script AI Agent",               "Phase 2 — LLM reads screen, reasons, acts; no test code needed"),
        ("#3", "LLM Reasoning Audit Trail",          "Phase 5 — per-step human-readable narrative from the AI"),
        ("#4", "Admin Approval Workflow",             "Phase 7 — report gated behind super-admin sign-off"),
    ]
    for j, (num, title, desc) in enumerate(core_items):
        r = 5 + j
        ws.row_dimensions[r].height = 20
        bg = "FFF2CC" if j % 2 == 0 else "FFEEBA"
        for col, val in enumerate([num, f"{title}  —  {desc}"], 2):
            cell = ws.cell(row=r, column=col, value=val)
            cell.fill = PatternFill("solid", fgColor=bg)
            cell.font = Font(name="Calibri", bold=(col==2), size=9,
                             color=("C00000" if col==2 else "333333"))
            cell.alignment = Alignment(horizontal=("center" if col==2 else "left"),
                                        vertical="center", indent=(1 if col==3 else 0))
            cell.border = thin_border()
        ws.merge_cells(f"C{r}:F{r}")

    # ── Discipline breakdown ──────────────────────────────────────────────────
    ws.row_dimensions[10].height = 22
    ws.merge_cells("B10:F10")
    dh = ws["B10"]
    dh.value = "Effort by Discipline"
    dh.font = Font(name="Calibri", bold=True, size=11, color="FFFFFF")
    dh.fill = PatternFill("solid", fgColor="2F5496")
    dh.alignment = Alignment(horizontal="center", vertical="center")
    dh.border = thin_border()

    ws.row_dimensions[11].height = 22
    for col, header in enumerate(["Discipline", "Total Hours", "% of Build", "Equiv. Work Days (8 h)"], 2):
        cell = ws.cell(row=11, column=col, value=header)
        cell.fill = PatternFill("solid", fgColor="44546A")
        cell.font = Font(name="Calibri", bold=True, size=10, color="FFFFFF")
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = thin_border()

    disc_data = [
        ("Frontend",  all_fe,  "2E75B6"),
        ("Backend",   all_be,  "375623"),
        ("QA",        all_qa,  "833C00"),
        ("DevOps",    all_dv,  "44546A"),
        ("Database",  all_db,  "7030A0"),
        ("TOTAL",     all_tot, "1F3864"),
    ]
    for i, (disc, hrs, color) in enumerate(disc_data):
        r = 12 + i
        ws.row_dimensions[r].height = 22
        is_total = (disc == "TOTAL")
        bg = color if is_total else "F2F2F2"
        fg = "FFFFFF" if is_total else color

        for col, (val, fmt) in enumerate([(disc, None), (hrs, '0 "hrs"'),
                                          (hrs/all_tot if all_tot else 0, '0.0%'),
                                          (hrs/8, '0.0 "days"')], 2):
            cell = ws.cell(row=r, column=col, value=val)
            cell.fill = PatternFill("solid", fgColor=(color if col==2 else bg))
            cell.font = Font(name="Calibri", bold=is_total, size=10,
                             color=("FFFFFF" if col==2 else fg))
            cell.alignment = Alignment(horizontal=("left" if col==2 else "center"),
                                       vertical="center", indent=(1 if col==2 else 0))
            cell.border = thin_border()
            if fmt:
                cell.number_format = fmt

    # ── Phase breakdown table ─────────────────────────────────────────────────
    ph_start = 21
    ws.row_dimensions[ph_start].height = 22
    ws.merge_cells(f"B{ph_start}:F{ph_start}")
    ph_hdr = ws[f"B{ph_start}"]
    ph_hdr.value = "Hours by Phase"
    ph_hdr.font = Font(name="Calibri", bold=True, size=11, color="FFFFFF")
    ph_hdr.fill = PatternFill("solid", fgColor="2F5496")
    ph_hdr.alignment = Alignment(horizontal="center", vertical="center")
    ph_hdr.border = thin_border()

    ws.row_dimensions[ph_start+1].height = 20
    for col, h in enumerate(["Phase", "FE", "BE", "QA", "DevOps / DB", "Total"], 2):
        c = ws.cell(row=ph_start+1, column=col, value=h)
        c.fill = PatternFill("solid", fgColor="44546A")
        c.font = Font(name="Calibri", bold=True, size=9, color="FFFFFF")
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border = thin_border()

    phases = {}
    for (id_, phase, feature, sub, tag, fe, be, qa, dv, db) in WBS:
        if phase not in phases:
            phases[phase] = [0,0,0,0,0]
        phases[phase][0] += fe
        phases[phase][1] += be
        phases[phase][2] += qa
        phases[phase][3] += dv
        phases[phase][4] += db

    alt = False
    r = ph_start + 2
    for phase, (fe, be, qa, dv, db) in phases.items():
        ws.row_dimensions[r].height = 18
        tot = fe + be + qa + dv + db
        bg = "EAF2FB" if not alt else "FFFFFF"
        color = PHASE_COLORS.get(phase, "1F3864")
        alt = not alt

        for i, val in enumerate([phase, fe, be, qa, dv+db, tot]):
            cell = ws.cell(row=r, column=2+i, value=val)
            cell.fill = PatternFill("solid", fgColor=(color if i==0 else bg))
            cell.font = Font(name="Calibri", size=9, bold=(i in (0,5)),
                             color=("FFFFFF" if i==0 else "333333"))
            cell.alignment = Alignment(horizontal=("left" if i==0 else "center"),
                                       vertical="center", indent=(1 if i==0 else 0))
            cell.border = thin_border()
        r += 1

    ws.row_dimensions[r].height = 22
    for i, (val, col_color) in enumerate(zip(
            ["TOTAL", all_fe, all_be, all_qa, all_dv+all_db, all_tot],
            ["1F3864","2E75B6","375623","833C00","44546A","1F3864"])):
        cell = ws.cell(row=r, column=2+i, value=val)
        cell.fill = PatternFill("solid", fgColor=col_color)
        cell.font = Font(name="Calibri", bold=True, size=10, color="FFFFFF")
        cell.alignment = Alignment(horizontal=("left" if i==0 else "center"),
                                   vertical="center", indent=(1 if i==0 else 0))
        cell.border = header_border()

    # ── Assumptions ───────────────────────────────────────────────────────────
    note_row = r + 3
    ws.merge_cells(f"B{note_row}:F{note_row}")
    note_hdr = ws.cell(row=note_row, column=2, value="Estimation Assumptions")
    note_hdr.font = Font(name="Calibri", bold=True, size=10, color="FFFFFF")
    note_hdr.fill = PatternFill("solid", fgColor="1F3864")
    note_hdr.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    note_hdr.border = thin_border()
    ws.row_dimensions[note_row].height = 20

    notes = [
        "• Hours = single skilled developer per discipline; not parallelised headcount.",
        "• 1 work day = 8 productive hours.  Recommend +20% contingency on grand total.",
        "• Phase D + A (Design) hours are Front-End effort; UX designer time is included in FE column.",
        "• ★★★ Core Diff rows are the 4 differentiators that set traindroid apart from all competitors.",
        "• ★★ MVP rows are required for Milestone 1 working prototype (Week 1 target).",
        "• ★ M2+ rows are Milestone 2 or later; exclude from Week-1 scope.",
        "• Database column: schema design, migrations, index tuning only — no DBA licence cost.",
        "• AI API costs (per-run) are runtime costs, not build effort — see PITCH.md for breakdown.",
        "• Emulator Docker spike (Phase A.6, 12h DevOps) may reduce once AVD-in-Docker is proven.",
    ]
    for j, note in enumerate(notes):
        nr = note_row + 1 + j
        ws.merge_cells(f"B{nr}:F{nr}")
        cell = ws.cell(row=nr, column=2, value=note)
        cell.font = Font(name="Calibri", size=9, color="444444")
        cell.fill = PatternFill("solid", fgColor=("F7F9FC" if j%2==0 else "FFFFFF"))
        cell.alignment = Alignment(horizontal="left", vertical="center", indent=1, wrap_text=True)
        cell.border = thin_border()
        ws.row_dimensions[nr].height = 16


def build_milestone_sheet(wb):
    ws = wb.create_sheet("Milestone Plan")
    ws.sheet_view.showGridLines = False

    col_widths = [4, 22, 46, 11, 11, 11, 11, 11, 11, 4]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    ws.merge_cells("A1:J1")
    t = ws["A1"]
    t.value = "traindroid — Milestone-Scoped Effort Plan"
    t.font = Font(name="Calibri", bold=True, size=14, color="FFFFFF")
    t.fill = PatternFill("solid", fgColor="1F3864")
    t.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 28

    milestones = [
        {
            "name": "Milestone 0 — Design Sprint (Days 1–3)",
            "color": "6E2594",
            "light": "F3E5F5",
            "goal": "Wireframes, architecture diagram, API contracts, emulator Docker spike, LLM prompt PoC",
            "items": [
                ("User Research & Personas",                          4,  0,  1,  0,  0),
                ("Wireframes — all 6 portal screens",                 8,  0,  1,  0,  0),
                ("Hi-Fi Mockups + Design System tokens",             12,  0,  1,  0,  0),
                ("System Architecture Diagram (C4)",                  0,  4,  0,  2,  0),
                ("API Contract Design (OpenAPI spec)",                0,  6,  1,  0,  0),
                ("Database ER Diagram",                               0,  2,  0,  0,  4),
                ("ADR: LLM provider, DB, queue, Docker",              0,  4,  0,  0,  0),
                ("Tech Spike — Emulator in Docker (PoC)",             0,  0,  1, 12,  0),
                ("Tech Spike — LLM Action Schema (PoC)",              0,  6,  1,  0,  0),
                ("Security Threat Model",                             0,  3,  2,  0,  0),
            ],
        },
        {
            "name": "Milestone 1 — MVP (Week 1)",
            "color": "1D6A96",
            "light": "D6EAF8",
            "goal": "Core loop: APK upload → AI agent tests → LLM reasoning audit log → pass/fail report",
            "items": [
                ("Repo, Docker Compose, .env, DB schema",             0,  6,  1,  8,  6),
                ("APK upload endpoint + self-hosted volume [★★★]",    0,  4,  2,  0,  2),
                ("File validation + security checks [★★★]",           0,  3,  2,  0,  0),
                ("Android Emulator Docker Image + ADB layer [★★★]",   0,  6,  2, 12,  0),
                ("Screenshot capture + UI tree dump [★★★]",           0,  7,  2,  0,  0),
                ("LLM prompt engineering + Tier 1 integration [★★★]", 0, 14,  3,  0,  0),
                ("Agent loop orchestrator [★★★]",                     0, 10,  3,  0,  0),
                ("Action execution via ADB [★★★]",                    0,  6,  2,  0,  0),
                ("Per-step audit log + LLM reasoning narrative [★★★]",0, 10,  3,  0,  4),
                ("Audit trail immutability [★★★]",                    0,  2,  1,  0,  2),
                ("HTML + PDF report generation",                      3,  7,  2,  0,  0),
                ("Portal: App shell, upload screen, config screen",  14,  4,  3,  0,  0),
                ("Portal: Live progress + results/audit screen",     18,  6,  4,  0,  0),
                ("Portal: Admin approval screen + API [★★★]",         7,  6,  3,  0,  2),
                ("Stuck-screen detection + max-steps guard",          0,  4,  2,  0,  0),
                ("Structured logging + error tracking",               0,  4,  1,  2,  0),
                ("Unit + integration tests (backend)",                0,  4,  8,  0,  0),
                ("Security test pass (OWASP upload / auth)",          0,  2,  6,  0,  0),
                ("Developer setup guide + API docs",                  2,  5,  1,  1,  0),
            ],
        },
        {
            "name": "Milestone 2 — Hardening (Weeks 2–3)",
            "color": "375623",
            "light": "D5F5E3",
            "goal": "Vision fallback, configurable goals, notifications, load/perf testing, Docker production build",
            "items": [
                ("Tier 2 vision model + tiered routing logic",        0, 10,  3,  0,  0),
                ("Confidence threshold + human review flag",          0,  4,  2,  0,  0),
                ("Multi-flow / configurable goal support",            0,  6,  2,  0,  0),
                ("Signup + Login reference flow (built-in)",          0,  4,  2,  0,  0),
                ("Emulator pool manager (slot allocation)",           0,  6,  2,  2,  0),
                ("Admin email notification on pending runs",          0,  3,  1,  0,  0),
                ("AI cost tracking per run",                          0,  3,  1,  0,  2),
                ("Design system implementation + UI polish",          8,  0,  2,  0,  0),
                ("Loading, error & empty states",                     4,  0,  2,  0,  0),
                ("E2E Playwright tests",                              2,  0, 10,  0,  0),
                ("Performance + load testing",                        0,  0,  6,  2,  0),
                ("Docker multi-stage production build",               0,  0,  1,  6,  0),
                ("Metrics dashboard (Grafana + Prometheus)",          0,  0,  1,  6,  2),
                ("Architecture overview doc + CONTRIBUTING.md",       0,  5,  0,  1,  0),
            ],
        },
        {
            "name": "Milestone 3 — Scale (Month 2)",
            "color": "843C0C",
            "light": "FDEBD0",
            "goal": "CI/CD integration, concurrent runs, RBAC, regression comparison, Slack notifications",
            "items": [
                ("GitHub Actions workflow + HMAC webhook trigger",    0,  6,  2,  8,  0),
                ("Status badge + completion callback",                0,  3,  1,  2,  0),
                ("Multiple concurrent test runs (queue)",             0,  8,  3,  2,  0),
                ("Regression comparison (baseline vs current)",       0,  8,  4,  0,  2),
                ("Role-based access control (admin / tester)",        4,  6,  2,  0,  2),
                ("Slack / email notification system",                 0,  6,  2,  0,  0),
                ("Accessibility (a11y) implementation",               4,  0,  3,  0,  0),
                ("LLM API cost monitoring dashboard",                 0,  2,  1,  0,  2),
                ("User portal guide + deployment guide",              2,  3,  1,  2,  0),
            ],
        },
    ]

    row = 3
    grand = [0, 0, 0, 0, 0]

    for ms in milestones:
        ws.merge_cells(f"A{row}:J{row}")
        mh = ws.cell(row=row, column=1, value=f"  {ms['name']}")
        mh.font = Font(name="Calibri", bold=True, size=12, color="FFFFFF")
        mh.fill = PatternFill("solid", fgColor=ms["color"])
        mh.alignment = Alignment(vertical="center")
        mh.border = thin_border()
        ws.row_dimensions[row].height = 22
        row += 1

        ws.merge_cells(f"B{row}:J{row}")
        gc = ws.cell(row=row, column=2, value=f"Goal: {ms['goal']}")
        gc.font = Font(name="Calibri", italic=True, size=9, color="555555")
        gc.fill = PatternFill("solid", fgColor=ms["light"])
        gc.alignment = Alignment(vertical="center", indent=1, wrap_text=True)
        gc.border = thin_border()
        ws.row_dimensions[row].height = 22
        row += 1

        ws.row_dimensions[row].height = 18
        for col, h in enumerate(["#", "Deliverable", "", "FE", "BE", "QA", "DevOps", "DB", "Total", ""], 1):
            ch = ws.cell(row=row, column=col, value=h)
            ch.fill = PatternFill("solid", fgColor=ms["color"])
            ch.font = Font(name="Calibri", bold=True, size=9, color="FFFFFF")
            ch.alignment = Alignment(horizontal="center", vertical="center")
            ch.border = thin_border()
        row += 1

        ms_totals = [0, 0, 0, 0, 0]
        for idx, (desc, fe, be, qa, dv, db) in enumerate(ms["items"]):
            tot = fe + be + qa + dv + db
            bg = ms["light"] if idx % 2 == 0 else "FFFFFF"
            vals = [idx+1, desc, "", fe, be, qa, dv, db, tot, ""]
            for col, val in enumerate(vals, 1):
                cell = ws.cell(row=row, column=col, value=val)
                cell.fill = PatternFill("solid", fgColor=bg)
                cell.font = Font(name="Calibri", size=9, color="333333", bold=(col==9))
                cell.alignment = Alignment(
                    horizontal=("center" if col in (1,4,5,6,7,8,9) else "left"),
                    vertical="center", wrap_text=(col==2))
                cell.border = thin_border()
            ws.row_dimensions[row].height = 18
            for k, v in enumerate([fe, be, qa, dv, db]):
                ms_totals[k] += v
                grand[k] += v
            row += 1

        ms_tot = sum(ms_totals)
        ws.merge_cells(f"A{row}:C{row}")
        sc = ws.cell(row=row, column=1, value=f"  {ms['name']} Subtotal")
        sc.font = Font(name="Calibri", bold=True, size=9, color="FFFFFF")
        sc.fill = PatternFill("solid", fgColor=ms["color"])
        sc.border = thin_border()
        sc.alignment = Alignment(vertical="center")
        for i, val in enumerate(ms_totals + [ms_tot]):
            cell = ws.cell(row=row, column=4+i, value=val)
            cell.fill = PatternFill("solid", fgColor=ms["color"])
            cell.font = Font(name="Calibri", bold=True, size=10, color="FFFFFF")
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = header_border()
        ws.row_dimensions[row].height = 20
        row += 2

    gt_tot = sum(grand)
    ws.merge_cells(f"A{row}:C{row}")
    gt = ws.cell(row=row, column=1, value="  GRAND TOTAL (all milestones)")
    gt.font = Font(name="Calibri", bold=True, size=11, color="FFFFFF")
    gt.fill = PatternFill("solid", fgColor="1F3864")
    gt.border = header_border()
    gt.alignment = Alignment(vertical="center")
    colors_gt = ["2E75B6","375623","833C00","44546A","7030A0","1F3864"]
    for i, (val, col_color) in enumerate(zip(grand + [gt_tot], colors_gt)):
        cell = ws.cell(row=row, column=4+i, value=val)
        cell.fill = PatternFill("solid", fgColor=col_color)
        cell.font = Font(name="Calibri", bold=True, size=11, color="FFFFFF")
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = header_border()
    ws.row_dimensions[row].height = 26


# ──────────────────────────────────────────────────────────────────────────────
# MAIN
# ──────────────────────────────────────────────────────────────────────────────

def main():
    wb = Workbook()
    wb, all_fe, all_be, all_qa, all_dv, all_db, all_tot = build_wbs_sheet(wb)
    build_summary_sheet(wb, all_fe, all_be, all_qa, all_dv, all_db, all_tot)
    build_milestone_sheet(wb)

    out = Path(__file__).parent / "traindroid_WBS_v2.xlsx"
    wb.save(out)

    print(f"\nSaved: {out}")
    print(f"\n{'─'*60}")
    print(f"  traindroid WBS v2.0 — Discipline Totals")
    print(f"{'─'*60}")
    print(f"  Frontend  : {all_fe:>5} hrs  ({all_fe/8:5.1f} days)")
    print(f"  Backend   : {all_be:>5} hrs  ({all_be/8:5.1f} days)")
    print(f"  QA        : {all_qa:>5} hrs  ({all_qa/8:5.1f} days)")
    print(f"  DevOps    : {all_dv:>5} hrs  ({all_dv/8:5.1f} days)")
    print(f"  Database  : {all_db:>5} hrs  ({all_db/8:5.1f} days)")
    print(f"{'─'*60}")
    print(f"  TOTAL     : {all_tot:>5} hrs  ({all_tot/8:5.1f} days)")
    print(f"  +20% buf  : {int(all_tot*1.2):>5} hrs  ({all_tot*1.2/8:5.1f} days)")
    print(f"{'─'*60}")

    # Core diff breakdown
    core_fe  = sum(r[5] for r in WBS if r[4]==T_CORE)
    core_be  = sum(r[6] for r in WBS if r[4]==T_CORE)
    core_qa  = sum(r[7] for r in WBS if r[4]==T_CORE)
    core_dv  = sum(r[8] for r in WBS if r[4]==T_CORE)
    core_db  = sum(r[9] for r in WBS if r[4]==T_CORE)
    core_tot = core_fe + core_be + core_qa + core_dv + core_db
    print(f"\n  ★★★ Core Differentiator hours: {core_tot} hrs ({core_tot/8:.1f} days)")
    print(f"  ★★  MVP (non-core) hours      : {sum(r[5]+r[6]+r[7]+r[8]+r[9] for r in WBS if r[4]==T_MVP):>4} hrs")
    print(f"  ★   M2+ hours                 : {sum(r[5]+r[6]+r[7]+r[8]+r[9] for r in WBS if r[4]==T_M2):>4} hrs")
    print(f"  —   Design hours              : {sum(r[5]+r[6]+r[7]+r[8]+r[9] for r in WBS if r[4]==T_DSGN):>4} hrs")
    print(f"{'─'*60}\n")


if __name__ == "__main__":
    main()
